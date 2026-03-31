"""Module: scripts/sf_insert_via_api.py

Salesforce Data Insert Script for VQMS (via REST API).

Connects to Salesforce using credentials from .env and inserts
all vendor data (Accounts, Contacts, Contracts) directly via
the simple-salesforce Python library. No CSV upload or manual
Salesforce UI interaction needed.

Prerequisites:
  1. Salesforce Developer Org with custom fields created (see plan Step 2)
  2. .env file with Salesforce credentials filled in (see plan Step 6)
  3. Python packages: simple-salesforce, python-dotenv, openpyxl

Insert order (parent objects first):
  1. Accounts  (25 records) — vendors
  2. Contacts  (50 records) — linked to Accounts via Vendor_ID__c
  3. Contracts (25 records) — linked to Accounts via Vendor_ID__c

Usage:
  uv run python scripts/sf_insert_via_api.py
  # or
  python scripts/sf_insert_via_api.py

Options:
  --dry-run    Show what would be inserted without actually inserting
  --skip-accounts   Skip Account insertion (if already imported)
  --skip-contacts   Skip Contact insertion (if already imported)
  --skip-contracts  Skip Contract insertion (if already imported)
"""

from __future__ import annotations

import os
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install with: uv add openpyxl")
    sys.exit(1)

try:
    from dotenv import load_dotenv
except ImportError:
    print("ERROR: python-dotenv is required. Install with: uv add python-dotenv")
    sys.exit(1)

try:
    from simple_salesforce import Salesforce, SalesforceAuthenticationFailed
except ImportError:
    print("ERROR: simple-salesforce is required. Install with: uv add simple-salesforce")
    sys.exit(1)


# --- Configuration ---

EXCEL_FILE_PATH = Path(
    "C:/Users/ROHIT/Downloads/Spreadsheets/2026-February/VQMS_Dummy_Dataset.xlsx"
)


# --- Helper Functions ---

def strip_currency(value: str | None) -> float:
    """Remove currency symbols and commas, return as float.

    '₹5,000,000' -> 5000000.0
    """
    if value is None:
        return 0.0
    text = re.sub(r"[₹$,\s]", "", str(value))
    try:
        return float(text)
    except ValueError:
        return 0.0


def split_full_name(full_name: str | None) -> tuple[str, str]:
    """Split 'Sneha Singh' into ('Sneha', 'Singh')."""
    if not full_name:
        return ("", "")
    name = str(full_name).strip()
    if " " in name:
        parts = name.rsplit(" ", 1)
        return (parts[0], parts[1])
    return ("", name)


def parse_location(location: str | None) -> tuple[str, str]:
    """Parse 'Pune, Maharashtra' into ('Pune', 'Maharashtra')."""
    if not location:
        return ("", "")
    text = str(location).strip()
    if "," in text:
        parts = text.split(",", 1)
        return (parts[0].strip(), parts[1].strip())
    return (text, "")


def format_date(value: str | None) -> str | None:
    """Format date to YYYY-MM-DD string. Returns None if empty."""
    if value is None:
        return None
    text = str(value).strip()
    if " " in text:
        text = text.split(" ")[0]
    if re.match(r"\d{4}-\d{2}-\d{2}", text):
        return text[:10]
    return None


def read_sheet_as_dicts(workbook: openpyxl.Workbook, sheet_name: str) -> list[dict]:
    """Read an Excel sheet and return a list of dicts."""
    ws = workbook[sheet_name]
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        rows.append(dict(zip(headers, row)))
    return rows


# --- Salesforce Insert Functions ---

def insert_accounts(
    sf: Salesforce,
    vendors: list[dict],
    *,
    dry_run: bool = False,
) -> dict[str, str]:
    """Insert vendor records as Salesforce Accounts.

    Returns a mapping of vendor_id -> Salesforce Account Id
    for linking Contacts and Contracts.
    """
    print("\n" + "=" * 60)
    print("STEP 1: Inserting Accounts (Vendors)")
    print("=" * 60)

    vendor_to_sf_id: dict[str, str] = {}
    success_count = 0
    error_count = 0

    for i, vendor in enumerate(vendors, 1):
        city, state = parse_location(vendor.get("location"))
        domain = str(vendor.get("domain", "") or "").strip()
        website = f"https://{domain}" if domain and not domain.startswith("http") else domain

        account_data = {
            "Name": vendor.get("company_name", ""),
            "Vendor_ID__c": vendor.get("vendor_id", ""),
            "Website": website,
            "Vendor_Tier__c": vendor.get("vendor_tier", ""),
            "Category__c": str(vendor.get("category", "") or ""),
            "Payment_Terms__c": vendor.get("payment_terms", ""),
            "AnnualRevenue": strip_currency(vendor.get("annual_contract_value")),
            "SLA_Response_Hours__c": vendor.get("sla_response_hours"),
            "SLA_Resolution_Days__c": vendor.get("sla_resolution_days"),
            "Vendor_Status__c": vendor.get("status", ""),
            "Onboarded_Date__c": format_date(vendor.get("onboarded_date")),
            "BillingCity": city,
            "BillingState": state,
            "BillingCountry": "India",
        }

        vid = vendor.get("vendor_id", "")

        if dry_run:
            print(f"  [{i:>2}/{len(vendors)}] DRY RUN: Would insert Account '{account_data['Name']}' ({vid})")
            vendor_to_sf_id[vid] = f"FAKE-ID-{vid}"
            success_count += 1
            continue

        try:
            result = sf.Account.create(account_data)
            sf_id = result["id"]
            vendor_to_sf_id[vid] = sf_id
            success_count += 1
            print(f"  [{i:>2}/{len(vendors)}] OK: {account_data['Name']} ({vid}) -> {sf_id}")
        except Exception as e:
            error_count += 1
            print(f"  [{i:>2}/{len(vendors)}] ERROR: {account_data['Name']} ({vid}) -> {e}")

    print(f"\n  Results: {success_count} inserted, {error_count} errors")
    return vendor_to_sf_id


def insert_contacts(
    sf: Salesforce,
    contacts: list[dict],
    vendor_to_sf_id: dict[str, str],
    *,
    dry_run: bool = False,
) -> None:
    """Insert vendor contact records as Salesforce Contacts.

    Links each Contact to its parent Account using the
    vendor_to_sf_id mapping from insert_accounts().
    """
    print("\n" + "=" * 60)
    print("STEP 2: Inserting Contacts (Vendor Contacts)")
    print("=" * 60)

    success_count = 0
    error_count = 0
    skip_count = 0

    for i, contact in enumerate(contacts, 1):
        first_name, last_name = split_full_name(contact.get("full_name"))
        vid = contact.get("vendor_id", "")
        account_id = vendor_to_sf_id.get(vid)

        if not account_id:
            skip_count += 1
            print(f"  [{i:>2}/{len(contacts)}] SKIP: No Account found for vendor_id={vid}")
            continue

        is_active = str(contact.get("is_active", "true")).strip().lower() == "true"

        contact_data = {
            "AccountId": account_id,
            "Contact_ID__c": contact.get("contact_id", ""),
            "FirstName": first_name,
            "LastName": last_name,
            "Email": contact.get("email", ""),
            "Phone": contact.get("phone", ""),
            "Title": contact.get("role", ""),
            "Contact_Type__c": contact.get("contact_type", ""),
            "Is_Active__c": is_active,
        }

        cid = contact.get("contact_id", "")

        if dry_run:
            print(f"  [{i:>2}/{len(contacts)}] DRY RUN: Would insert Contact '{first_name} {last_name}' ({cid}) -> Account {vid}")
            success_count += 1
            continue

        try:
            result = sf.Contact.create(contact_data)
            success_count += 1
            print(f"  [{i:>2}/{len(contacts)}] OK: {first_name} {last_name} ({cid}) -> {result['id']}")
        except Exception as e:
            error_count += 1
            print(f"  [{i:>2}/{len(contacts)}] ERROR: {first_name} {last_name} ({cid}) -> {e}")

    print(f"\n  Results: {success_count} inserted, {error_count} errors, {skip_count} skipped")


def insert_contracts(
    sf: Salesforce,
    contracts: list[dict],
    vendor_to_sf_id: dict[str, str],
    *,
    dry_run: bool = False,
) -> None:
    """Insert contract records as Salesforce Contracts.

    Links each Contract to its parent Account. All contracts
    are inserted with Status='Draft' because Salesforce requires
    Contracts to start as Draft.
    """
    print("\n" + "=" * 60)
    print("STEP 3: Inserting Contracts")
    print("=" * 60)

    success_count = 0
    error_count = 0
    skip_count = 0

    for i, contract in enumerate(contracts, 1):
        vid = contract.get("vendor_id", "")
        account_id = vendor_to_sf_id.get(vid)

        if not account_id:
            skip_count += 1
            print(f"  [{i:>2}/{len(contracts)}] SKIP: No Account found for vendor_id={vid}")
            continue

        contract_data = {
            "AccountId": account_id,
            "Contract_ID__c": contract.get("contract_id", ""),
            "StartDate": format_date(contract.get("start_date")),
            "EndDate": format_date(contract.get("end_date")),
            "Status": "Draft",
            "Payment_Terms__c": contract.get("payment_terms", ""),
            "Contract_Value__c": strip_currency(contract.get("contract_value")),
            "SLA_Response_Hours__c": contract.get("sla_response_hrs"),
            "SLA_Resolution_Days__c": contract.get("sla_resolution_days"),
            "Late_Penalty__c": str(contract.get("late_penalty", "") or ""),
            "Review_Frequency__c": contract.get("review_frequency", ""),
            "Description": str(contract.get("notes", "") or ""),
        }

        ctid = contract.get("contract_id", "")

        if dry_run:
            print(f"  [{i:>2}/{len(contracts)}] DRY RUN: Would insert Contract '{ctid}' -> Account {vid}")
            success_count += 1
            continue

        try:
            result = sf.Contract.create(contract_data)
            success_count += 1
            print(f"  [{i:>2}/{len(contracts)}] OK: {ctid} -> {result['id']}")
        except Exception as e:
            error_count += 1
            print(f"  [{i:>2}/{len(contracts)}] ERROR: {ctid} -> {e}")

    print(f"\n  Results: {success_count} inserted, {error_count} errors, {skip_count} skipped")


def lookup_existing_accounts(sf: Salesforce) -> dict[str, str]:
    """Query Salesforce for existing vendor Accounts.

    Used when --skip-accounts is passed to get the Account IDs
    for linking Contacts and Contracts.
    """
    print("\n  Looking up existing Accounts...")
    result = sf.query(
        "SELECT Id, Vendor_ID__c FROM Account WHERE Vendor_ID__c LIKE 'V-%'"
    )
    mapping = {}
    for record in result["records"]:
        mapping[record["Vendor_ID__c"]] = record["Id"]
    print(f"  Found {len(mapping)} existing vendor Accounts")
    return mapping


# --- Main Entry Point ---

def main() -> None:
    """Read Excel data and insert into Salesforce via REST API."""
    # Parse command-line flags
    args = set(sys.argv[1:])
    dry_run = "--dry-run" in args
    skip_accounts = "--skip-accounts" in args
    skip_contacts = "--skip-contacts" in args
    skip_contracts_flag = "--skip-contracts" in args

    # Load .env
    load_dotenv()

    # Validate credentials
    username = os.getenv("SALESFORCE_USERNAME", "")
    password = os.getenv("SALESFORCE_PASSWORD", "")
    security_token = os.getenv("SALESFORCE_SECURITY_TOKEN", "")

    missing = []
    if not username:
        missing.append("SALESFORCE_USERNAME")
    if not password:
        missing.append("SALESFORCE_PASSWORD")
    if not security_token:
        missing.append("SALESFORCE_SECURITY_TOKEN")

    if missing and not dry_run:
        print("ERROR: Missing Salesforce credentials in .env:")
        for var in missing:
            print(f"  - {var}")
        print("\nEither fill in .env or use --dry-run to preview without connecting.")
        sys.exit(1)

    # Verify Excel file
    if not EXCEL_FILE_PATH.exists():
        print(f"ERROR: Excel file not found at: {EXCEL_FILE_PATH}")
        sys.exit(1)

    # Read Excel data
    print("=" * 60)
    print("VQMS SALESFORCE DATA IMPORT (via Python API)")
    print("=" * 60)
    if dry_run:
        print("\n  *** DRY RUN MODE — no records will be inserted ***\n")

    print(f"  Reading: {EXCEL_FILE_PATH}")
    wb = openpyxl.load_workbook(EXCEL_FILE_PATH, read_only=True)

    vendors = read_sheet_as_dicts(wb, "vendors")
    contacts = read_sheet_as_dicts(wb, "vendor_contacts")
    contracts_data = read_sheet_as_dicts(wb, "contracts")

    print(f"  Vendors:  {len(vendors)} rows")
    print(f"  Contacts: {len(contacts)} rows")
    print(f"  Contracts: {len(contracts_data)} rows")

    # Connect to Salesforce
    sf = None
    if not dry_run:
        print(f"\n  Connecting to Salesforce as {username}...")
        try:
            sf = Salesforce(
                username=username,
                password=password,
                security_token=security_token,
            )
            print(f"  Connected! API URL: {sf.base_url}")
        except SalesforceAuthenticationFailed as e:
            print(f"\n  ERROR: Authentication failed: {e}")
            print("  Check your .env credentials.")
            sys.exit(1)

    # Insert in order: Accounts -> Contacts -> Contracts
    vendor_to_sf_id: dict[str, str] = {}

    # Step 1: Accounts
    if skip_accounts:
        print("\n  Skipping Account insertion (--skip-accounts)")
        if sf:
            vendor_to_sf_id = lookup_existing_accounts(sf)
        else:
            # Dry run with skip — create fake IDs
            for v in vendors:
                vid = v.get("vendor_id", "")
                vendor_to_sf_id[vid] = f"FAKE-ID-{vid}"
    else:
        vendor_to_sf_id = insert_accounts(sf, vendors, dry_run=dry_run)

    # Step 2: Contacts
    if skip_contacts:
        print("\n  Skipping Contact insertion (--skip-contacts)")
    else:
        insert_contacts(sf, contacts, vendor_to_sf_id, dry_run=dry_run)

    # Step 3: Contracts
    if skip_contracts_flag:
        print("\n  Skipping Contract insertion (--skip-contracts)")
    else:
        insert_contracts(sf, contracts_data, vendor_to_sf_id, dry_run=dry_run)

    wb.close()

    # Summary
    print("\n" + "=" * 60)
    if dry_run:
        print("DRY RUN COMPLETE — no records were inserted")
        print("Remove --dry-run to insert for real")
    else:
        print("IMPORT COMPLETE")
        print(f"\nVerify in Salesforce:")
        print(f"  Accounts:  Setup -> Accounts tab -> All Accounts")
        print(f"  Contacts:  Setup -> Contacts tab -> All Contacts")
        print(f"  Contracts: Setup -> Contracts tab")
        print(f"\nOr run: uv run python scripts/test_sf_connection.py")
    print("=" * 60)


if __name__ == "__main__":
    main()
