"""Module: scripts/prepare_sf_import.py

Salesforce Import CSV Preparation Script for VQMS.

Reads the VQMS_Dummy_Dataset.xlsx file and transforms three sheets
(vendors, vendor_contacts, contracts) into Salesforce-ready CSV files.

Transformations applied:
  - Currency symbols (₹) and commas stripped from numeric fields
  - Full names split into FirstName / LastName
  - Location "City, State" parsed into BillingCity / BillingState
  - Domain prefixed with https://
  - Contract status mapped: ACTIVE -> Draft (SF requirement)
  - Dates formatted as YYYY-MM-DD

Output files (written to scripts/sf_import_data/):
  - sf_accounts.csv   (25 rows -> Salesforce Account object)
  - sf_contacts.csv   (50 rows -> Salesforce Contact object)
  - sf_contracts.csv  (25 rows -> Salesforce Contract object)

Usage:
  uv run python scripts/prepare_sf_import.py
  # or
  python scripts/prepare_sf_import.py
"""

from __future__ import annotations

import csv
import os
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install with: uv add openpyxl")
    print("  or: pip install openpyxl")
    sys.exit(1)


# --- Configuration ---

# Path to the Excel file — update this if the file is in a different location
EXCEL_FILE_PATH = Path(
    "C:/Users/ROHIT/Downloads/Spreadsheets/2026-February/VQMS_Dummy_Dataset.xlsx"
)

# Output directory for generated CSVs
OUTPUT_DIR = Path(__file__).parent / "sf_import_data"


# --- Helper Functions ---

def strip_currency(value: str | None) -> str:
    """Remove currency symbols (₹, $) and commas from a string, returning a plain number.

    Examples:
        '₹5,000,000' -> '5000000'
        '$10,800,000' -> '10800000'
        'None' or None -> ''
    """
    if value is None:
        return ""
    text = str(value)
    # Remove currency symbols, commas, and whitespace
    cleaned = re.sub(r"[₹$,\s]", "", text)
    # If nothing numeric remains, return empty
    if not cleaned or not any(c.isdigit() for c in cleaned):
        return ""
    return cleaned


def split_full_name(full_name: str | None) -> tuple[str, str]:
    """Split a full name into (first_name, last_name).

    Splits on the LAST space so names like 'Sneha Singh' become
    ('Sneha', 'Singh') and 'Ravi Kumar Patel' becomes ('Ravi Kumar', 'Patel').

    If there's no space, the entire name goes into last_name
    (Salesforce requires LastName but FirstName is optional).
    """
    if not full_name:
        return ("", "")
    name = str(full_name).strip()
    if " " in name:
        # Split on last space to handle multi-word first names
        parts = name.rsplit(" ", 1)
        return (parts[0], parts[1])
    # Single name — Salesforce requires LastName
    return ("", name)


def parse_location(location: str | None) -> tuple[str, str]:
    """Parse 'City, State' into (city, state).

    Examples:
        'Pune, Maharashtra' -> ('Pune', 'Maharashtra')
        'Mumbai, Maharashtra' -> ('Mumbai', 'Maharashtra')
        None -> ('', '')
    """
    if not location:
        return ("", "")
    text = str(location).strip()
    if "," in text:
        parts = text.split(",", 1)
        return (parts[0].strip(), parts[1].strip())
    return (text, "")


def format_date(value: str | None) -> str:
    """Format a date value to YYYY-MM-DD string.

    Handles both string dates and datetime objects from openpyxl.
    """
    if value is None:
        return ""
    text = str(value).strip()
    # If it's a datetime string like '2022-02-04 00:00:00', take just the date part
    if " " in text:
        text = text.split(" ")[0]
    # If it looks like a date (YYYY-MM-DD), return as-is
    if re.match(r"\d{4}-\d{2}-\d{2}", text):
        return text[:10]
    return text


def read_sheet_as_dicts(workbook: openpyxl.Workbook, sheet_name: str) -> list[dict]:
    """Read an Excel sheet and return a list of dicts (one per row).

    The first row is treated as headers. Skips completely empty rows.
    """
    ws = workbook[sheet_name]
    headers = [cell.value for cell in ws[1]]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Skip rows where all values are None
        if all(v is None for v in row):
            continue
        row_dict = dict(zip(headers, row))
        rows.append(row_dict)

    return rows


# --- CSV Generation Functions ---

def generate_accounts_csv(vendors: list[dict], output_path: Path) -> int:
    """Transform vendor rows into Salesforce Account CSV.

    Returns the number of rows written.
    """
    fieldnames = [
        "Vendor_ID__c",
        "Name",
        "Website",
        "Vendor_Tier__c",
        "Category__c",
        "Payment_Terms__c",
        "AnnualRevenue",
        "SLA_Response_Hours__c",
        "SLA_Resolution_Days__c",
        "Vendor_Status__c",
        "Onboarded_Date__c",
        "BillingCity",
        "BillingState",
        "BillingCountry",
    ]

    rows_written = 0
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()

        for vendor in vendors:
            city, state = parse_location(vendor.get("location"))
            domain = str(vendor.get("domain", "")).strip()
            # Prefix domain with https:// if not already present
            website = f"https://{domain}" if domain and not domain.startswith("http") else domain

            row = {
                "Vendor_ID__c": vendor.get("vendor_id", ""),
                "Name": vendor.get("company_name", ""),
                "Website": website,
                "Vendor_Tier__c": vendor.get("vendor_tier", ""),
                "Category__c": vendor.get("category", ""),
                "Payment_Terms__c": vendor.get("payment_terms", ""),
                "AnnualRevenue": strip_currency(vendor.get("annual_contract_value")),
                "SLA_Response_Hours__c": vendor.get("sla_response_hours", ""),
                "SLA_Resolution_Days__c": vendor.get("sla_resolution_days", ""),
                "Vendor_Status__c": vendor.get("status", ""),
                "Onboarded_Date__c": format_date(vendor.get("onboarded_date")),
                "BillingCity": city,
                "BillingState": state,
                "BillingCountry": "India",
            }
            writer.writerow(row)
            rows_written += 1

    return rows_written


def generate_contacts_csv(contacts: list[dict], output_path: Path) -> int:
    """Transform vendor_contacts rows into Salesforce Contact CSV.

    Uses Account.Vendor_ID__c to link each Contact to its parent Account
    via External ID lookup during import.

    Returns the number of rows written.
    """
    fieldnames = [
        "Contact_ID__c",
        "Account.Vendor_ID__c",
        "FirstName",
        "LastName",
        "Email",
        "Phone",
        "Title",
        "Contact_Type__c",
        "Is_Active__c",
    ]

    rows_written = 0
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()

        for contact in contacts:
            first_name, last_name = split_full_name(contact.get("full_name"))
            # Convert is_active string to lowercase for Salesforce checkbox
            is_active = str(contact.get("is_active", "true")).strip().lower()

            row = {
                "Contact_ID__c": contact.get("contact_id", ""),
                "Account.Vendor_ID__c": contact.get("vendor_id", ""),
                "FirstName": first_name,
                "LastName": last_name,
                "Email": contact.get("email", ""),
                "Phone": contact.get("phone", ""),
                "Title": contact.get("role", ""),
                "Contact_Type__c": contact.get("contact_type", ""),
                "Is_Active__c": is_active,
            }
            writer.writerow(row)
            rows_written += 1

    return rows_written


def generate_contracts_csv(contracts: list[dict], output_path: Path) -> int:
    """Transform contracts rows into Salesforce Contract CSV.

    Uses Account.Vendor_ID__c to link each Contract to its parent Account.
    Maps status ACTIVE -> Draft because Salesforce requires Contracts
    to start as Draft (they can be activated later via the UI or API).

    Returns the number of rows written.
    """
    fieldnames = [
        "Contract_ID__c",
        "Account.Vendor_ID__c",
        "StartDate",
        "EndDate",
        "Status",
        "Payment_Terms__c",
        "Contract_Value__c",
        "SLA_Response_Hours__c",
        "SLA_Resolution_Days__c",
        "Late_Penalty__c",
        "Review_Frequency__c",
        "Description",
    ]

    rows_written = 0
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()

        for contract in contracts:
            # Salesforce Contract must start as "Draft" — cannot insert as "Activated"
            # We map ACTIVE -> Draft here; activate later in SF if needed
            status = "Draft"

            row = {
                "Contract_ID__c": contract.get("contract_id", ""),
                "Account.Vendor_ID__c": contract.get("vendor_id", ""),
                "StartDate": format_date(contract.get("start_date")),
                "EndDate": format_date(contract.get("end_date")),
                "Status": status,
                "Payment_Terms__c": contract.get("payment_terms", ""),
                "Contract_Value__c": strip_currency(contract.get("contract_value")),
                "SLA_Response_Hours__c": contract.get("sla_response_hrs", ""),
                "SLA_Resolution_Days__c": contract.get("sla_resolution_days", ""),
                "Late_Penalty__c": contract.get("late_penalty", ""),
                "Review_Frequency__c": contract.get("review_frequency", ""),
                "Description": contract.get("notes", ""),
            }
            writer.writerow(row)
            rows_written += 1

    return rows_written


# --- Main Entry Point ---

def main() -> None:
    """Read the Excel file, transform data, and write 3 Salesforce-ready CSVs."""
    # Verify Excel file exists
    if not EXCEL_FILE_PATH.exists():
        print(f"ERROR: Excel file not found at: {EXCEL_FILE_PATH}")
        print("Update EXCEL_FILE_PATH in this script to point to your file.")
        sys.exit(1)

    print(f"Reading Excel file: {EXCEL_FILE_PATH}")
    wb = openpyxl.load_workbook(EXCEL_FILE_PATH, read_only=True)

    # Read the three sheets we need
    vendors = read_sheet_as_dicts(wb, "vendors")
    contacts = read_sheet_as_dicts(wb, "vendor_contacts")
    contracts = read_sheet_as_dicts(wb, "contracts")

    print(f"  vendors:         {len(vendors)} rows")
    print(f"  vendor_contacts: {len(contacts)} rows")
    print(f"  contracts:       {len(contracts)} rows")

    # Create output directory
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    print(f"\nOutput directory: {OUTPUT_DIR}")

    # Generate CSVs
    accounts_path = OUTPUT_DIR / "sf_accounts.csv"
    contacts_path = OUTPUT_DIR / "sf_contacts.csv"
    contracts_path = OUTPUT_DIR / "sf_contracts.csv"

    n_accounts = generate_accounts_csv(vendors, accounts_path)
    print(f"\n  sf_accounts.csv:  {n_accounts} rows written")

    n_contacts = generate_contacts_csv(contacts, contacts_path)
    print(f"  sf_contacts.csv:  {n_contacts} rows written")

    n_contracts = generate_contracts_csv(contracts, contracts_path)
    print(f"  sf_contracts.csv: {n_contracts} rows written")

    wb.close()

    # Print summary
    print("\n" + "=" * 60)
    print("SALESFORCE IMPORT FILES READY")
    print("=" * 60)
    print(f"\n  1. {accounts_path}")
    print(f"     -> Import into Salesforce ACCOUNT object ({n_accounts} records)")
    print(f"\n  2. {contacts_path}")
    print(f"     -> Import into Salesforce CONTACT object ({n_contacts} records)")
    print(f"     -> Links to Accounts via Vendor_ID__c external ID")
    print(f"\n  3. {contracts_path}")
    print(f"     -> Import into Salesforce CONTRACT object ({n_contracts} records)")
    print(f"     -> Links to Accounts via Vendor_ID__c external ID")
    print(f"\nIMPORT ORDER: Accounts FIRST, then Contacts, then Contracts")
    print("=" * 60)


if __name__ == "__main__":
    main()
